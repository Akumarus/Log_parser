from operator import index
from openpyxl import Workbook
import pandas as pd
import os
from pandas import DataFrame, ExcelFile, ExcelWriter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

class CsvHandler:

    @staticmethod
    def load_csv(path: str) -> pd.DataFrame:
        
        if path is None:
            print("Путь к файлу не установлен.")
            return None
        
        if os.path.exists(path):
            print(f"Путь {path} существует.")
            """Загружает данные из CSV файла в DataFrame"""
            df = pd.read_csv(path, sep=';', names=["Время", "ID"] + [f"Байт_{i}" for i in range(8)])
            #print(df)

            # Удаляет все строки с NaN
            df = df.dropna() 
            for col in df.select_dtypes(include=['float64']):
                df[col] = df[col].apply(lambda x: int(x) if isinstance(x, float) and not pd.isna(x) and x.is_integer() else x)
        
            return df
        else:
            print(f"Путь {path} не существует.")
            return None
    
    @staticmethod
    def save_csv(tables):
        """Сохраняет DataFrame в CSV файл"""
        script_path = os.path.dirname(os.path.abspath(__file__))
        folder_path = os.path.join(script_path, "saved_tables")
        os.makedirs(folder_path, exist_ok=True)
        
        for id, table in tables.items():
            print(f"id: {id}")
            #print(f"table {table}")
            #file_path = os.path.join(folder_path, f"table id{id}.csv")
            #table.to_csv(file_path, encoding='utf-8-sig', index=False, sep=';')
            
            # Сохраняем таблицу в Excel
            #file_path = os.path.join(folder_path, f"table id{id}.xlsx")
            #with ExcelWriter(file_path, engine='openpyxl') as writer:
            #    table.to_excel(writer, index=False, sheet_name=f"Table_{id}")
            #print(f"Сохранена таблица для ID {id} в файл {file_path}")
            
            # Убедимся, что файл будет с расширением .xlsx
            file_path = os.path.join(folder_path, f"table_id_{id}.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = f"Table_{id}"

             # Добавляем данные из DataFrame в лист Excel
            for r_idx, row in enumerate(dataframe_to_rows(table, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    # Устанавливаем жирный шрифт для заголовков
                    if r_idx == 1:
                        cell.font = Font(bold=True)

            for col_idx, column_name in enumerate(table.columns, start=1):
                max_length = max(len(str(column_name)), *(len(str(val)) for val in table[column_name]))
                adjusted_width = max_length + 2  # Добавляем небольшой запас
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width

            wb.save(file_path)
